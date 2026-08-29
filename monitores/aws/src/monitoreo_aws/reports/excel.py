from __future__ import annotations
from pathlib import Path
from datetime import datetime, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
import re
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY='E36C0A'; BLUE='F28C28'; SKY='FFF3E6'; TEAL='78BE20'; BG='FFF9F4'; WHITE='FFFFFF'; GREEN='5A9E1F'; RED='C62828'; AMBER='F6A53A'; SOFT_RED='FDECEC'; SOFT_GREEN='EEF7E8'; SOFT_AMBER='FFF4E6'; SOFT_BLUE='FFF5EB'; GRID='EADCCF'; DARK='5B402A'; GREY='8A765F'
THIN_SIDE = Side(style='thin', color=GRID)
THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _bogota_tz():
    try:
        return ZoneInfo("America/Bogota")
    except ZoneInfoNotFoundError:
        from datetime import timedelta, timezone as dt_timezone
        return dt_timezone(timedelta(hours=-5), name="America/Bogota")


def _local_datetime(value, hour_only=False):
    text = str(value or "").strip()
    if not text or not re.search(r"\d{4}-\d{2}-\d{2}", text):
        return text
    normalized = text.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(normalized)
    except ValueError:
        match = re.search(r"(\d{4}-\d{2}-\d{2})[T\s](\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)", text)
        if not match:
            return text
        try:
            dt = datetime.fromisoformat(f"{match.group(1)}T{match.group(2)}")
        except ValueError:
            return text
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    dt = dt.astimezone(_bogota_tz())
    return dt.strftime("%H:%M" if hour_only else "%Y-%m-%d %H:%M:%S")


def _num(value) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _fill(color):
    return PatternFill('solid', fgColor=color)


def _base_style(cell, fill=None, bold=False, color=DARK, align='left'):
    cell.border = THIN
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.font = Font(name='Segoe UI', size=10, bold=bold, color=color)
    if fill:
        cell.fill = _fill(fill)


def _page_title(ws, text, cols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, text)
    c.fill = _fill(NAVY)
    c.font = Font(name='Segoe UI', size=16, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28


def _info_bar(ws, row, values, cols=8):
    col = 1
    for label, value in values:
        ws.cell(row, col, label)
        _base_style(ws.cell(row, col), fill=SOFT_BLUE, bold=True)
        ws.merge_cells(start_row=row, start_column=col+1, end_row=row, end_column=col+2)
        ws.cell(row, col+1, value)
        _base_style(ws.cell(row, col+1), fill=WHITE)
        col += 3
        if col > cols:
            break


def _section_title(ws, row, text, cols=8, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.fill = _fill(color)
    c.font = Font(name='Segoe UI', size=12, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = THIN
    ws.row_dimensions[row].height = 22


def _subsection_title(ws, row, text, cols=8, color=SOFT_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.fill = _fill(color)
    c.font = Font(name='Segoe UI', size=10, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = THIN


def _header(ws, row, headers, fill=NAVY):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = _fill(fill)
        c.font = Font(name='Segoe UI', size=10, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN
    ws.row_dimensions[row].height = 20


def _data_row(ws, row, values, fill=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row, col, v)
        _base_style(c, fill=fill)
    return row + 1


def _set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.sheet_view.showGridLines = False


def _status_generic(value, alert_limit=None, no_data_alert=False):
    if value is None:
        return 'SIN DATO', GREY, SOFT_BLUE
    n = _num(value)
    if no_data_alert and n == 0:
        return 'SIN INFORMACIÓN', RED, SOFT_RED
    if alert_limit is not None and n >= alert_limit:
        return 'ALERTA', RED, SOFT_RED
    return 'NORMALIDAD', GREEN, SOFT_GREEN


def _status_tup(value):
    if value is None:
        return 'SIN DATO', GREY, SOFT_BLUE, 'No fue posible consultar el log.'
    n = _num(value)
    if n <= 30:
        return 'NORMAL', GREEN, SOFT_GREEN, 'Dentro del nivel normal (0 a 30).'
    if n <= 200:
        return 'REGULAR', AMBER, SOFT_AMBER, 'Fuera de lo normal, pero dentro del rango regular (31 a 200).'
    if n <= 250:
        return 'ATENCIÓN', AMBER, SOFT_AMBER, 'Nivel alto de atención (201 a 250).'
    return 'CRÍTICA', RED, SOFT_RED, 'Nivel preocupante o crítico (más de 250).'


def _paint_status(cell, text, color, fill):
    cell.value = text
    _base_style(cell, fill=fill, bold=True, color=color, align='center')


def _top_peak(rows):
    if not rows:
        return '', 0
    peak = max(rows, key=lambda r: _num(r.get('count')))
    return _local_datetime(peak.get('hora', ''), hour_only=True), _num(peak.get('count'))


def _group_by_hour(*series_lists):
    bucket = {}
    for idx, rows in enumerate(series_lists):
        for r in rows:
            hour = _local_datetime(r.get('hora', ''), hour_only=True)
            bucket.setdefault(hour, [0] * len(series_lists))
            bucket[hour][idx] = _num(r.get('count'))
    return sorted(bucket.items())


def _build_summary(ws, cfg, ventana, data, alertas):
    m = data['metricas']
    _page_title(ws, 'MONITOREO AWS · RESUMEN EJECUTIVO · COMPENSAR', 8)
    _info_bar(ws, 2, [('Analista', cfg['app']['analista']), ('Corte', ventana.nombre)])
    _info_bar(ws, 3, [('Inicio', str(ventana.inicio)), ('Fin', str(ventana.fin))])
    _section_title(ws, 5, 'RESUMEN GENERAL', 8)
    _header(ws, 6, ['Grupo', 'Indicador', 'Valor', 'Estado', 'Observación', 'Detalle', 'Inicio', 'Fin'])
    pago_ok = sum(_num(m.get(k)) for k in ('aprob_creacion_payu','aprob_creacion_ecollect','aprob_estado_payu','aprob_estado_ecollect','aprob_receiver'))
    pago_err = sum(_num(m.get(k)) for k in ('err_creacion_payu','err_creacion_ecollect','err_estado_payu','err_estado_ecollect','err_receiver'))
    mens_err = _num(m.get('mens_timeout')) + _num(m.get('mens_503')) + _num(m.get('mens_502')) + _num(m.get('mens_cannot')) + _num(m.get('mens_sms_failed')) + _num(m.get('mens_error_400_total')) + _num(m.get('otp_408'))
    summary_rows = [
        ('API PAGOS', 'Aprobadas totales', pago_ok, 'NORMALIDAD', 'Vista consolidada de aprobaciones', 'Incluye creación, estado y receiver', str(ventana.inicio), str(ventana.fin)),
        ('API PAGOS', 'Errores totales', pago_err, 'ALERTA' if pago_err >= 41 else 'NORMALIDAD', 'Errores operativos del orquestador', 'Incluye PayU, Ecollect y Receiver', str(ventana.inicio), str(ventana.fin)),
        ('CSC', 'Errores proxy', _num(m.get('csc_504')) + _num(m.get('csc_task_timed')), 'NORMALIDAD', 'Errores 504 + Task timed', 'Lámbda de proxy de pagos', str(ventana.inicio), str(ventana.fin)),
        ('MENSAJERÍA', 'Transaccionalidad', _num(m.get('mens_total_send')), 'NORMALIDAD', 'Total REPORT del corte', 'No incluye request manuales', str(ventana.inicio), str(ventana.fin)),
        ('MENSAJERÍA', 'Errores totales', mens_err, 'ALERTA' if any(a['grupo']=='MENSAJERÍA' for a in alertas) else 'NORMALIDAD', 'Timeout, 502, 503, 400, 408, cannot, SMS', 'Ver hoja MENSAJERÍA', str(ventana.inicio), str(ventana.fin)),
        ('TARJETA TUP', 'Errores TUP', _num(m.get('tup_error')) if m.get('tup_error') is not None else 'N/D', _status_tup(m.get('tup_error'))[0], _status_tup(m.get('tup_error'))[3], 'Ver detalle por hora en hoja TARJETA TUP', str(ventana.inicio), str(ventana.fin)),
        ('API SUBSIDIOS', 'Error creación archivo CX', _num(m.get('error_cx')) if m.get('error_cx') is not None else 'N/D', 'ALERTA' if _num(m.get('error_cx')) > 0 else 'NORMALIDAD', 'Control del proceso CX', 'Ver hoja API SUBSIDIOS', str(ventana.inicio), str(ventana.fin)),
    ]
    row = 7
    for r in summary_rows:
        row = _data_row(ws, row, list(r), fill=WHITE if row % 2 else BG)
    _section_title(ws, row + 1, 'TRANSACCIONALIDAD GENERAL', 8, color=TEAL)
    row += 2
    _header(ws, row, ['Servicio', 'Resultado', 'Estado', 'Observación', 'Servicio', 'Resultado', 'Estado', 'Observación'], fill=TEAL)
    row += 1
    trans_left = [
        ('API Mensajería (REPORT)', _num(m.get('mens_total_send')), 'REGISTRADO', 'Total sent del corte'),
        ('Replicador corporativo', _num(m.get('replicador')), 'REGISTRADO' if _num(m.get('replicador')) > 0 else 'ALERTA', 'Debe replicar al menos 1 vez'),
        ('Validar OTP 500', _num(m.get('otp_500')), 'NORMALIDAD' if _num(m.get('otp_500')) == 0 else 'ALERTA', 'Errores 500 de validar OTP'),
    ]
    trans_right = [
        ('API Módulo Seguridad', _num(m.get('seg_consulta_persona')), 'REGISTRADO', 'Consulta Persona / seguridad'),
        ('Errores generales Tarjeta TUP', _num(m.get('tup_error')), _status_tup(m.get('tup_error'))[0], 'Errores funcionales TUP'),
        ('Error creación archivo CX', _num(m.get('error_cx')), 'ALERTA' if _num(m.get('error_cx')) > 0 else 'NORMALIDAD', 'API Subsidios'),
    ]
    for left, right in zip(trans_left, trans_right):
        values = list(left) + list(right)
        row = _data_row(ws, row, values, fill=SOFT_BLUE if row % 2 else WHITE)
    chart_start = row + 2
    _header(ws, chart_start, ['Indicador', 'Valor'])
    chart_data = [
        ('Pagos aprobados', pago_ok), ('Pagos con error', pago_err), ('Mensajería enviados', _num(m.get('mens_total_send'))),
        ('Mensajería errores', mens_err), ('Replicador', _num(m.get('replicador'))), ('TUP errores', _num(m.get('tup_error'))),
    ]
    for item in chart_data:
        chart_start += 1
        _data_row(ws, chart_start, item)
    bar = BarChart()
    bar.title = 'Volumen consolidado del corte'
    bar.y_axis.title = 'Cantidad'
    bar.height = 7
    bar.width = 14
    bar.style = 11
    bar.add_data(Reference(ws, min_col=2, min_row=chart_start-len(chart_data)+1, max_row=chart_start), titles_from_data=False)
    bar.set_categories(Reference(ws, min_col=1, min_row=chart_start-len(chart_data)+1, max_row=chart_start))
    ws.add_chart(bar, 'J6')
    _set_widths(ws, [18, 24, 14, 16, 35, 38, 22, 22])
    ws.freeze_panes = 'A6'
    return ws


def _build_api_pagos(ws, ventana, data):
    m = data['metricas']
    _page_title(ws, 'API PAGOS · DETALLE DEL CORTE · COMPENSAR', 6)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Grupo', 'INTEROPPROD')], cols=6)
    row = 4
    _section_title(ws, row, 'TRANSACCIONES APROBADAS / 200', 6)
    row += 1
    _header(ws, row, ['Proceso', 'Proveedor', 'Resultado', 'Estado', 'Inicio', 'Fin'])
    approved = [
        ('Creación de pago', 'PayU', m.get('aprob_creacion_payu')),
        ('Creación de pago', 'Ecollect', m.get('aprob_creacion_ecollect')),
        ('Estado de transacciones', 'PayU', m.get('aprob_estado_payu')),
        ('Estado de transacciones', 'Ecollect', m.get('aprob_estado_ecollect')),
        ('Notificación al bus', 'Receiver', m.get('aprob_receiver')),
    ]
    for proc, prov, val in approved:
        st, color, fill = _status_generic(val, no_data_alert=True)
        row += 1
        _data_row(ws, row, [proc, prov, 'N/D' if val is None else _num(val), '', str(ventana.inicio), str(ventana.fin)], fill=WHITE if row % 2 else BG)
        _paint_status(ws.cell(row, 4), st, color, fill)
    row += 2
    _section_title(ws, row, 'ERRORES', 6, color=RED)
    row += 1
    _header(ws, row, ['Proceso', 'Proveedor', 'Resultado', 'Estado', 'Inicio', 'Fin'], fill=RED)
    errors = [
        ('Creación de pago', 'PayU', m.get('err_creacion_payu')),
        ('Creación de pago', 'Ecollect', m.get('err_creacion_ecollect')),
        ('Estado de transacciones', 'PayU', m.get('err_estado_payu')),
        ('Estado de transacciones', 'Ecollect', m.get('err_estado_ecollect')),
        ('Notificación al bus', 'Receiver', m.get('err_receiver')),
    ]
    for proc, prov, val in errors:
        st, color, fill = _status_generic(val, alert_limit=41)
        row += 1
        _data_row(ws, row, [proc, prov, 'N/D' if val is None else _num(val), '', str(ventana.inicio), str(ventana.fin)], fill=WHITE if row % 2 else BG)
        _paint_status(ws.cell(row, 4), st, color, fill)
    row += 2
    _section_title(ws, row, 'COMPARATIVO RÁPIDO', 6, color=TEAL)
    row += 1
    _header(ws, row, ['Indicador', 'Valor'])
    chart_rows_start = row + 1
    chart_data = [
        ('Aprobadas creación', _num(m.get('aprob_creacion_payu')) + _num(m.get('aprob_creacion_ecollect'))),
        ('Aprobadas estado', _num(m.get('aprob_estado_payu')) + _num(m.get('aprob_estado_ecollect'))),
        ('Aprobadas receiver', _num(m.get('aprob_receiver'))),
        ('Errores creación', _num(m.get('err_creacion_payu')) + _num(m.get('err_creacion_ecollect'))),
        ('Errores estado', _num(m.get('err_estado_payu')) + _num(m.get('err_estado_ecollect'))),
        ('Errores receiver', _num(m.get('err_receiver'))),
    ]
    for item in chart_data:
        row += 1
        _data_row(ws, row, item)
    chart = BarChart(); chart.title = 'API Pagos: aprobadas vs errores'; chart.y_axis.title = 'Cantidad'; chart.height = 7; chart.width = 13; chart.style = 10
    chart.add_data(Reference(ws, min_col=2, min_row=chart_rows_start, max_row=row), titles_from_data=False)
    chart.set_categories(Reference(ws, min_col=1, min_row=chart_rows_start, max_row=row)); ws.add_chart(chart, 'H5')
    _set_widths(ws, [26, 18, 14, 16, 22, 22]); ws.freeze_panes = 'A5'; return ws

def _build_csc(ws, ventana, data):
    m = data['metricas']
    _page_title(ws, 'CSC · DETALLE DEL CORTE · COMPENSAR', 7)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Grupo', 'CSC')], cols=7)
    row = 4
    _section_title(ws, row, 'ERRORES DEL PROXY DE PAGOS', 7)
    row += 1
    _header(ws, row, ['Componente', 'Métrica', 'Resultado', 'Estado', 'Observación', 'Inicio', 'Fin'])
    specs = [
        ('Lambda Proxy PaymentsPost', '504 Gateway Time-out', m.get('csc_504'), 'Monitoreo de time-out del proxy'),
        ('Lambda Proxy PaymentsPost', 'Task timed', m.get('csc_task_timed'), 'Monitoreo de task timed out'),
    ]
    for comp, metric, val, obs in specs:
        st, color, fill = _status_generic(val, alert_limit=1)
        row += 1
        _data_row(ws, row, [comp, metric, 'N/D' if val is None else _num(val), '', obs, str(ventana.inicio), str(ventana.fin)], fill=WHITE if row % 2 else BG)
        _paint_status(ws.cell(row, 4), st, color, fill)
    _set_widths(ws, [28, 24, 14, 15, 36, 22, 22])
    ws.freeze_panes = 'A5'
    return ws


def _build_mensajeria(ws, ventana, data):
    m = data['metricas']
    d = data['detalles']
    _page_title(ws, 'MENSAJERÍA · TRANSACCIONALIDAD Y DETALLE · COMPENSAR', 8)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Grupo', 'MENSAJERÍA')], cols=8)
    row = 4
    _section_title(ws, row, 'TRANSACCIONALIDAD', 8)
    row += 1
    _header(ws, row, ['Servicio', 'Resultado', 'Estado', 'Observación', 'Servicio', 'Resultado', 'Estado', 'Observación'])
    left = [
        ('API Mensajería (REPORT)', m.get('mens_total_send'), 'NORMALIDAD', 'Total enviados del corte'),
        ('Replicador corporativo', m.get('replicador'), 'NORMALIDAD' if _num(m.get('replicador')) > 0 else 'ALERTA', 'Debe replicar mínimo 1 vez'),
        ('Validar OTP 500', m.get('otp_500'), 'NORMALIDAD' if _num(m.get('otp_500')) == 0 else 'ALERTA', 'Error 500 de ValidarOTP'),
    ]
    right = [
        ('Exitosos 200', m.get('mens_exitos_200_total'), 'REGISTRADO', 'Consolidado de 200'),
        ('Errores 400', m.get('mens_error_400_total'), 'REGISTRADO', 'Consolidado de 400'),
        ('OTP 408', m.get('otp_408'), 'REGISTRADO', 'Errores 408 de OTP'),
    ]
    for a, b in zip(left, right):
        row += 1
        _data_row(ws, row, [a[0], 'N/D' if a[1] is None else _num(a[1]), a[2], a[3], b[0], 'N/D' if b[1] is None else _num(b[1]), b[2], b[3]], fill=WHITE if row % 2 else BG)
    row += 2
    _section_title(ws, row, 'ERRORES', 8, color=RED)
    row += 1
    _header(ws, row, ['Métrica', 'Resultado', 'Estado', 'Observación', 'Métrica', 'Resultado', 'Estado', 'Observación'], fill=RED)
    err_left = [
        ('Timeout', m.get('mens_timeout'), 'Monitoreo de timeout generales'),
        ('Error 503', m.get('mens_503'), 'Service Temporarily Unavailable'),
        ('Error 502', m.get('mens_502'), 'Bad Gateway'),
    ]
    err_right = [
        ('Error Cannot', m.get('mens_cannot'), 'Broker SD / Httpcode != 200'),
        ('SMS failed', m.get('mens_sms_failed'), 'Falla de envío SMS'),
        ('Error 400', m.get('mens_error_400_total'), 'Mayor a 100 del mismo dato es crítica'),
    ]
    for a, b in zip(err_left, err_right):
        row += 1
        _data_row(ws, row, [a[0], 'N/D' if a[1] is None else _num(a[1]), 'REGISTRADO', a[2], b[0], 'N/D' if b[1] is None else _num(b[1]), 'REGISTRADO', b[2]], fill=WHITE if row % 2 else BG)
    row += 2
    _section_title(ws, row, 'DETALLE DE ERRORES', 8, color=AMBER)
    row += 1
    _header(ws, row, ['IdConsumer', 'Broker', 'HTTP', 'Operación', 'Tipología', 'Cantidad', 'Desde', 'Hasta'], fill=AMBER)
    for det in d.get('mensajeria_errores', [])[:18]:
        row += 1
        _data_row(ws, row, [det.get('IdConsumer',''), det.get('MessageIn.configS3.Broker',''), det.get('Httpcode',''), det.get('OperationInvokerName',''), det.get('MessageOut.error') or det.get('MessageOut',''), _num(det.get('count')), _local_datetime(det.get('desde','')), _local_datetime(det.get('hasta',''))], fill=WHITE if row % 2 else BG)
    row += 2
    _section_title(ws, row, 'DETALLE DE EXITOSOS 200', 8, color=TEAL)
    row += 1
    _header(ws, row, ['IdConsumer', 'Broker', 'HTTP', 'Operación', 'Cantidad', '', '', ''], fill=TEAL)
    for det in d.get('mensajeria_exitos', [])[:18]:
        row += 1
        _data_row(ws, row, [det.get('IdConsumer',''), det.get('MessageIn.configS3.Broker',''), det.get('Httpcode',''), det.get('OperationInvokerName',''), _num(det.get('count')), '', '', ''], fill=WHITE if row % 2 else BG)
    _set_widths(ws, [22, 14, 10, 22, 54, 12, 22, 22])
    ws.freeze_panes = 'A5'
    return ws


def _build_tup(ws, ventana, data):
    m = data['metricas']
    rows = data['detalles'].get('tup_por_hora', [])
    _page_title(ws, 'TARJETA TUP · DETALLE DEL CORTE · COMPENSAR', 7)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Grupo', 'TARJETA TUP')], cols=7)
    status, color, fill, obs = _status_tup(m.get('tup_error'))
    peak_hour, peak_count = _top_peak(rows)
    row = 4
    _section_title(ws, row, 'RESUMEN', 7)
    row += 1
    _header(ws, row, ['Métrica', 'Resultado', 'Estado', 'Observación', 'Hora pico', 'Valor pico', 'Rango'])
    row += 1
    _data_row(ws, row, ['Errores Tarjeta TUP', 'N/D' if m.get('tup_error') is None else _num(m.get('tup_error')), '', obs, peak_hour, peak_count, ventana.texto])
    _paint_status(ws.cell(row, 3), status, color, fill)
    row += 2
    _section_title(ws, row, 'DETALLE POR HORA', 7, color=TEAL)
    row += 1
    _header(ws, row, ['Hora', 'Cantidad', 'Estado', 'Observación', '', '', ''], fill=TEAL)
    table_start = row + 1
    for det in rows:
        row += 1
        c = _num(det.get('count'))
        st, _, _, ob = _status_tup(c)
        _data_row(ws, row, [_local_datetime(det.get('hora',''), hour_only=True), c, st, ob, '', '', ''], fill=WHITE if row % 2 else BG)
    if row >= table_start:
        chart = LineChart()
        chart.title = 'Errores TUP por hora'
        chart.y_axis.title = 'Cantidad'
        chart.height = 7
        chart.width = 13
        chart.style = 10
        chart.add_data(Reference(ws, min_col=2, min_row=table_start, max_row=row), titles_from_data=False)
        chart.set_categories(Reference(ws, min_col=1, min_row=table_start, max_row=row))
        ws.add_chart(chart, 'I5')
    _set_widths(ws, [22, 12, 14, 38, 16, 12, 20])
    ws.freeze_panes = 'A5'
    return ws


def _build_subsidios(ws, ventana, data):
    m = data['metricas']
    _page_title(ws, 'API SUBSIDIOS · DETALLE DEL CORTE · COMPENSAR', 6)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Grupo', 'API SUBSIDIOS')], cols=6)
    row = 4
    _section_title(ws, row, 'CONTROL DEL SERVICIO', 6)
    row += 1
    _header(ws, row, ['Servicio', 'Resultado', 'Estado', 'Inicio', 'Fin', 'Grupo'])
    row += 1
    _data_row(ws, row, ['API Subsidios - Error archivo CX', _num(m.get('error_cx')) if m.get('error_cx') is not None else 'N/D', 'ALERTA' if _num(m.get('error_cx')) > 0 else 'NORMALIDAD', str(ventana.inicio), str(ventana.fin), 'SUBSIDIOS'])
    _set_widths(ws, [34, 14, 16, 22, 22, 16]); ws.freeze_panes = 'A5'; return ws


def _build_seguridad(ws, ventana, data):
    m = data['metricas']
    _page_title(ws, 'MÓDULO DE SEGURIDAD · DETALLE DEL CORTE · COMPENSAR', 6)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Grupo', 'SEGURIDAD')], cols=6)
    row = 4
    _section_title(ws, row, 'TRANSACCIONALIDAD', 6, color=TEAL)
    row += 1
    _header(ws, row, ['Servicio', 'Resultado', 'Estado', 'Inicio', 'Fin', 'Grupo'], fill=TEAL)
    specs = [
        ('ConsultaPersona', m.get('seg_consulta_persona'), 'REGISTRADO' if m.get('seg_consulta_persona') is not None else 'SIN DATO'),
        ('Validar OTP 500', m.get('otp_500'), 'ALERTA' if _num(m.get('otp_500')) > 0 else 'NORMALIDAD'),
    ]
    for service, val, state in specs:
        row += 1
        _data_row(ws, row, [service, 'N/D' if val is None else _num(val), state, str(ventana.inicio), str(ventana.fin), 'SEGURIDAD'], fill=WHITE if row % 2 else BG)
    _set_widths(ws, [34, 14, 16, 22, 22, 16]); ws.freeze_panes = 'A5'; return ws

def _build_alerts(ws, ventana, alertas):
    _page_title(ws, 'ALERTAS DEL CORTE', 7)
    _info_bar(ws, 2, [('Rango', ventana.texto), ('Total alertas', str(len(alertas)))], cols=7)
    _header(ws, 4, ['Nivel', 'Grupo', 'Servicio', 'Métrica', 'Valor', 'Detalle', 'Rango'])
    row = 4
    if not alertas:
        row += 1
        _data_row(ws, row, ['OK', 'GENERAL', 'Todos', 'Sin alertas', 0, 'No se detectaron condiciones de alerta en este corte.', ventana.texto], fill=SOFT_GREEN)
    else:
        for a in alertas:
            row += 1
            fill = SOFT_RED if a['nivel'] == 'CRÍTICA' else SOFT_AMBER if a['nivel'] in ('ALTA', 'MEDIA') else SOFT_GREEN
            _data_row(ws, row, [a['nivel'], a['grupo'], a['servicio'], a['metrica'], a['valor'], a['detalle'], ventana.texto], fill=fill)
    _set_widths(ws, [12, 16, 24, 25, 12, 82, 34])
    ws.freeze_panes = 'A5'
    return ws


def _build_detalle_mens(ws, data):
    _page_title(ws, 'DETALLE DE MENSAJERÍA', 8)
    _header(ws, 3, ['IdConsumer', 'Broker', 'HTTP', 'Operación', 'Tipología', 'Cantidad', 'Desde', 'Hasta'])
    row = 3
    for det in data['detalles'].get('mensajeria_errores', []):
        row += 1
        _data_row(ws, row, [det.get('IdConsumer',''), det.get('MessageIn.configS3.Broker',''), det.get('Httpcode',''), det.get('OperationInvokerName',''), det.get('MessageOut.error') or det.get('MessageOut',''), _num(det.get('count')), _local_datetime(det.get('desde','')), _local_datetime(det.get('hasta',''))], fill=WHITE if row % 2 else BG)
    _set_widths(ws, [22, 14, 10, 24, 60, 12, 22, 22])
    ws.freeze_panes = 'A4'
    return ws


def _build_tendencias(ws, data):
    _page_title(ws, 'TENDENCIAS HORARIAS', 6)
    _header(ws, 3, ['Hora', 'Errores pagos', 'Errores TUP', 'Mensajería 400', 'Mensajería 200', 'Replicaciones'])
    row = 3
    grouped = _group_by_hour(data['detalles'].get('pagos_errores_por_hora', []), data['detalles'].get('tup_por_hora', []), data['detalles'].get('mensajeria_400_por_hora', []), data['detalles'].get('mensajeria_200_por_hora', []), data['detalles'].get('replicador_por_hora', []))
    start_data = row + 1
    for hour, vals in grouped:
        row += 1
        _data_row(ws, row, [hour, *vals], fill=WHITE if row % 2 else BG)
    if row >= start_data:
        chart = LineChart()
        chart.title = 'Tendencia horaria consolidada'
        chart.y_axis.title = 'Cantidad'
        chart.height = 8
        chart.width = 16
        chart.style = 10
        chart.add_data(Reference(ws, min_col=2, max_col=6, min_row=3, max_row=row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=row))
        ws.add_chart(chart, 'G4')
    _set_widths(ws, [24, 16, 16, 18, 18, 16])
    ws.freeze_panes = 'A4'
    return ws


def _build_metadata(ws, cfg, ventana, data):
    _page_title(ws, 'METADATOS DE EJECUCIÓN', 2)
    _header(ws, 3, ['Campo', 'Valor'])
    meta = [
        ('Corte', ventana.corte), ('Nombre', ventana.nombre), ('Inicio', str(ventana.inicio)), ('Fin', str(ventana.fin)),
        ('Zona horaria', cfg['app']['timezone']), ('Errores de consulta', len(data.get('errores_consulta', []))),
    ]
    row = 3
    for item in meta:
        row += 1
        _data_row(ws, row, item, fill=WHITE if row % 2 else BG)
    _set_widths(ws, [28, 86])
    ws.freeze_panes = 'A4'
    return ws


def generar_excel(path: Path, cfg: dict, ventana, data: dict, alertas: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = 'RESUMEN'
    _build_summary(ws, cfg, ventana, data, alertas)
    _build_api_pagos(wb.create_sheet('API PAGOS'), ventana, data)
    _build_csc(wb.create_sheet('CSC'), ventana, data)
    _build_mensajeria(wb.create_sheet('MENSAJERIA'), ventana, data)
    _build_tup(wb.create_sheet('TARJETA TUP'), ventana, data)
    _build_subsidios(wb.create_sheet('API SUBSIDIOS'), ventana, data)
    _build_seguridad(wb.create_sheet('SEGURIDAD'), ventana, data)
    _build_alerts(wb.create_sheet('ALERTAS'), ventana, alertas)
    _build_detalle_mens(wb.create_sheet('DETALLE MENSAJERIA'), data)
    _build_tendencias(wb.create_sheet('TENDENCIAS'), data)
    _build_metadata(wb.create_sheet('METADATOS'), cfg, ventana, data)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
